import streamlit as st
import pandas as pd
import camelot
from deep_translator import GoogleTranslator
from langdetect import detect
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# === Language Map ===
full_language_names = {
    "en": "English", "es": "Spanish", "fr": "French", "de": "German", "zh-cn": "Chinese (Simplified)",
    "zh-tw": "Chinese (Traditional)", "ja": "Japanese", "ko": "Korean", "ru": "Russian", "ar": "Arabic",
    "pt": "Portuguese", "it": "Italian", "hi": "Hindi", "tr": "Turkish", "pl": "Polish", "uk": "Ukrainian",
    "vi": "Vietnamese", "id": "Indonesian", "nl": "Dutch", "sv": "Swedish", "no": "Norwegian", "fi": "Finnish"
}
sorted_lang_names = sorted(full_language_names.values())
lang_name_to_code = {v: k for k, v in full_language_names.items()}

# === Helpers ===
def translate_text(text, target_lang):
    try:
        text = str(text).strip()
        if not text:
            return text
        if detect(text) != target_lang:
            return GoogleTranslator(source='auto', target=target_lang).translate(text)
    except:
        return text
    return text

def translate_df(df, target_lang):
    return df.applymap(lambda x: translate_text(x, target_lang))

def split_merged_rows(df):
    new_rows = []
    for _, row in df.iterrows():
        if any('\n' in str(cell) for cell in row):
            parts = [str(cell).split('\n') for cell in row]
            max_len = max(len(p) for p in parts)
            for i in range(max_len):
                new_row = [p[i] if i < len(p) else '' for p in parts]
                new_rows.append(new_row)
        else:
            new_rows.append(row.tolist())
    return pd.DataFrame(new_rows, columns=df.columns)

def format_excel(writer, sheet_name, df):
    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

def post_formatting(excel_io, sheet_names):
    wb = load_workbook(excel_io)
    for sheet in sheet_names:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 35
    wb.save(excel_io)

# === Streamlit UI ===
st.set_page_config(page_title="PDF Table Extractor + Translator", layout="centered")
st.title("📄 PDF Table Extractor + Translator")

# === Session State ===
if "last_filename" not in st.session_state:
    st.session_state.last_filename = None
if "confirmed" not in st.session_state:
    st.session_state.confirmed = False
if "processing_complete" not in st.session_state:
    st.session_state.processing_complete = False
if "confirmed_lang_code" not in st.session_state:
    st.session_state.confirmed_lang_code = "en"
    st.session_state.confirmed_lang_name = "English"
if "run_without_translation" not in st.session_state:
    st.session_state.run_without_translation = False

# === File Upload ===
new_file = st.file_uploader("Upload your PDF", type=["pdf"], key="uploaded_file")
if new_file and new_file.name != st.session_state.last_filename:
    st.session_state.last_filename = new_file.name
    st.session_state.confirmed = False
    st.session_state.processing_complete = False
    st.session_state.run_without_translation = False

enable_translation = False

# === Translation UI ===
if new_file:
    enable_translation = st.checkbox("🌍 Translate table content to another language?")

    if enable_translation:
        default_index = sorted_lang_names.index("English")
        selected_lang_name = st.selectbox("Choose target language:", sorted_lang_names, index=default_index, key="selected_lang")

        if st.button("✅ Confirm Language Selection"):
            st.session_state.confirmed_lang_code = lang_name_to_code[selected_lang_name]
            st.session_state.confirmed_lang_name = selected_lang_name
            st.session_state.confirmed = True
            st.session_state.processing_complete = False
            st.success(f"Language set to: {selected_lang_name}")

    elif not st.session_state.run_without_translation:
        if st.button("Extract Without Translation"):
            st.session_state.run_without_translation = True

# === PDF Processing ===
if new_file and (
    (enable_translation and st.session_state.confirmed) or
    (not enable_translation and st.session_state.run_without_translation)
):
    with st.spinner("Processing PDF..."):
        tables = camelot.read_pdf(new_file, pages='all', flavor='stream', strip_text='\n', edge_tol=200, row_tol=10)
        all_dfs = []
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')

        for i, table in enumerate(tables):
            df = table.df
            if df.empty:
                continue
            df.columns = [str(col).strip() for col in df.columns]
            df = split_merged_rows(df)
            df = df.astype(str)
            if enable_translation:
                df = translate_df(df, st.session_state.confirmed_lang_code)

            sheet_name = f"Page_{table.page}_Table_{i+1}"
            format_excel(writer, sheet_name, df)
            all_dfs.append(sheet_name[:31])

        writer.close()
        post_formatting(output, all_dfs)
        output.seek(0)

        st.success("✅ Done! Download your formatted Excel below:")
        st.download_button(
            label="📥 Download Translated Excel",
            data=output.getvalue(),
            file_name="translated_tables.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # ✅ Reset + rerun safely
        st.session_state.processing_complete = True
        st.session_state.confirmed = False
        st.session_state.run_without_translation = False
        st.query_params["refresh"] = str(pd.Timestamp.now())

# === Status Reminders ===
elif new_file and enable_translation and not st.session_state.confirmed:
    st.warning("Please confirm your language selection before processing.")
